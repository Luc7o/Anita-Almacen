import os, uuid, json
from utils.notificaciones import enviar_resumen_venta, enviar_alerta_stock
from functools import wraps
from datetime import datetime, timedelta
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, current_app
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from sqlalchemy import func
from app_extensions import db
from models.producto import Producto, Categoria, Proveedor
from models.movimiento import MovimientoStock
from models.venta import VentaFisica, DetalleVenta
from models.usuario import Usuario
from forms.almacen_forms import FormCategoria, FormProveedor, FormProducto

bp = Blueprint('admin', __name__)


def admin_requerido(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.es_admin:
            flash('Acceso denegado.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


def _guardar_imagen(archivo):
    # BUG CORREGIDO: Validar extensión antes de guardar para evitar subida de archivos maliciosos
    ext = archivo.filename.rsplit('.', 1)[-1].lower() if '.' in archivo.filename else ''
    if ext not in current_app.config.get('ALLOWED_EXTENSIONS', {'png', 'jpg', 'jpeg', 'webp'}):
        raise ValueError(f'Tipo de archivo no permitido: .{ext}')
    nombre = f"prod_{uuid.uuid4().hex[:12]}.{ext}"
    ruta   = os.path.join(current_app.config['UPLOAD_FOLDER'], nombre)
    archivo.save(ruta)
    return nombre


def _slugify(texto):
    # BUG CORREGIDO: Normalizar caracteres Unicode y eliminar dobles guiones
    reemplazos = {
        'á':'a','é':'e','í':'i','ó':'o','ú':'u','ñ':'n',
        'ä':'a','ë':'e','ï':'i','ö':'o','ü':'u',
        ' ':'-', '&':'-', '/':'-', '_':'-',
    }
    t = texto.lower()
    for k, v in reemplazos.items():
        t = t.replace(k, v)
    slug = ''.join(c for c in t if c.isalnum() or c == '-')
    # Eliminar guiones dobles o iniciales/finales
    import re
    slug = re.sub(r'-{2,}', '-', slug).strip('-')
    return slug


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/')
@admin_requerido
def dashboard():
    hoy   = datetime.utcnow().date()
    mes_i = datetime(hoy.year, hoy.month, 1)

    total_productos  = Producto.query.filter_by(activo=True).count()
    productos_sin_stock = Producto.query.filter(Producto.stock <= 0, Producto.activo==True).count()
    productos_stock_bajo = Producto.query.filter(
        Producto.stock > 0, Producto.stock <= Producto.stock_minimo, Producto.activo==True).count()

    ventas_hoy = db.session.query(func.sum(VentaFisica.total))\
        .filter(func.date(VentaFisica.fecha) == hoy, VentaFisica.anulada==False).scalar() or 0
    ventas_mes = db.session.query(func.sum(VentaFisica.total))\
        .filter(VentaFisica.fecha >= mes_i, VentaFisica.anulada==False).scalar() or 0
    num_ventas_hoy = VentaFisica.query\
        .filter(func.date(VentaFisica.fecha) == hoy, VentaFisica.anulada==False).count()

    # Valor total inventario
    valor_inventario = db.session.query(
        func.sum(Producto.precio_venta * Producto.stock)
    ).filter(Producto.activo==True).scalar() or 0

    # Últimos movimientos
    ultimos_movimientos = MovimientoStock.query\
        .order_by(MovimientoStock.fecha.desc()).limit(8).all()

    # Últimas ventas
    ultimas_ventas = VentaFisica.query\
        .filter_by(anulada=False)\
        .order_by(VentaFisica.fecha.desc()).limit(6).all()

    # Productos con stock bajo
    alertas_stock = Producto.query.filter(
        Producto.stock <= Producto.stock_minimo, Producto.activo==True
    ).order_by(Producto.stock.asc()).limit(8).all()

    # Ventas últimos 7 días para mini-chart
    ventas_semana = []
    for i in range(6, -1, -1):
        dia = hoy - timedelta(days=i)
        total = db.session.query(func.sum(VentaFisica.total))\
            .filter(func.date(VentaFisica.fecha) == dia, VentaFisica.anulada==False)\
            .scalar() or 0
        ventas_semana.append({'dia': dia.strftime('%d/%m'), 'total': float(total)})

    # Top productos más vendidos del mes
    top_productos = db.session.query(
        Producto.nombre,
        func.sum(DetalleVenta.cantidad).label('unidades'),
        func.sum(DetalleVenta.subtotal).label('monto')
    ).join(DetalleVenta, DetalleVenta.producto_id == Producto.id)\
     .join(VentaFisica, VentaFisica.id == DetalleVenta.venta_id)\
     .filter(VentaFisica.fecha >= mes_i, VentaFisica.anulada==False)\
     .group_by(Producto.id, Producto.nombre)\
     .order_by(func.sum(DetalleVenta.cantidad).desc())\
     .limit(5).all()

    return render_template('admin/dashboard.html',
        total_productos=total_productos,
        productos_sin_stock=productos_sin_stock,
        productos_stock_bajo=productos_stock_bajo,
        ventas_hoy=ventas_hoy, ventas_mes=ventas_mes,
        num_ventas_hoy=num_ventas_hoy,
        valor_inventario=valor_inventario,
        ultimos_movimientos=ultimos_movimientos,
        ultimas_ventas=ultimas_ventas,
        alertas_stock=alertas_stock,
        ventas_semana=ventas_semana,
        top_productos=top_productos,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# PRODUCTOS
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/productos')
@admin_requerido
def productos():
    pagina    = request.args.get('pagina', 1, type=int)
    busqueda  = request.args.get('q', '')
    cat_id    = request.args.get('cat', 0, type=int)
    estado    = request.args.get('estado', 'activos')

    query = Producto.query
    if estado == 'activos':
        query = query.filter_by(activo=True)
    elif estado == 'inactivos':
        query = query.filter_by(activo=False)
    if busqueda:
        query = query.filter(
            (Producto.nombre.ilike(f'%{busqueda}%')) |
            (Producto.sku.ilike(f'%{busqueda}%')) |
            (Producto.codigo_barras.ilike(f'%{busqueda}%'))
        )
    if cat_id:
        query = query.filter_by(categoria_id=cat_id)

    prods      = query.order_by(Producto.nombre.asc()).paginate(page=pagina, per_page=20)
    categorias = Categoria.query.filter_by(activo=True).all()
    return render_template('admin/productos.html', productos=prods,
                           busqueda=busqueda, categorias=categorias,
                           cat_sel=cat_id, estado=estado)


@bp.route('/productos/nuevo', methods=['GET', 'POST'])
@admin_requerido
def nuevo_producto():
    form = FormProducto()
    form.categoria_id.choices = [(c.id, c.nombre) for c in Categoria.query.filter_by(activo=True).all()]
    form.proveedor_id.choices = [(0, '— Sin proveedor —')] + [
        (p.id, p.nombre) for p in Proveedor.query.filter_by(activo=True).all()]
    if form.validate_on_submit():
        tallas = json.dumps([t.strip() for t in form.tallas.data.split(',') if t.strip()]) if form.tallas.data else None
        colores = json.dumps([c.strip() for c in form.colores.data.split(',') if c.strip()]) if form.colores.data else None
        prod = Producto(
            nombre=form.nombre.data, descripcion=form.descripcion.data,
            sku=form.sku.data or None, codigo_barras=form.codigo_barras.data or None,
            categoria_id=form.categoria_id.data,
            proveedor_id=form.proveedor_id.data if form.proveedor_id.data else None,
            precio_compra=form.precio_compra.data or 0,
            precio_venta=form.precio_venta.data,
            stock=form.stock.data or 0, stock_minimo=form.stock_minimo.data or 5,
            tallas=tallas, colores=colores, unidad=form.unidad.data,
            activo=form.activo.data,
        )
        if form.imagen.data and form.imagen.data.filename:
            try:
                prod.imagen_principal = _guardar_imagen(form.imagen.data)
            except ValueError as e:
                flash(str(e), 'danger')
                return render_template('admin/producto_form.html', form=form, titulo='Nuevo producto')
        db.session.add(prod)
        db.session.flush()
        # Registrar movimiento si hay stock inicial
        if prod.stock > 0:
            mov = MovimientoStock(
                producto_id=prod.id, tipo='entrada',
                cantidad=prod.stock, stock_antes=0, stock_despues=prod.stock,
                motivo='Stock inicial al registrar producto',
                usuario_id=current_user.id
            )
            db.session.add(mov)
        db.session.commit()
        flash('Producto creado exitosamente.', 'success')
        return redirect(url_for('admin.productos'))
    return render_template('admin/producto_form.html', form=form, titulo='Nuevo producto')


@bp.route('/productos/<int:id>/editar', methods=['GET', 'POST'])
@admin_requerido
def editar_producto(id):
    prod = Producto.query.get_or_404(id)
    form = FormProducto(obj=prod)
    form.categoria_id.choices = [(c.id, c.nombre) for c in Categoria.query.filter_by(activo=True).all()]
    form.proveedor_id.choices = [(0, '— Sin proveedor —')] + [
        (p.id, p.nombre) for p in Proveedor.query.filter_by(activo=True).all()]
    if request.method == 'GET':
        form.tallas.data  = ', '.join(prod.tallas_lista)
        form.colores.data = ', '.join(prod.colores_lista)
        form.proveedor_id.data = prod.proveedor_id or 0
    if form.validate_on_submit():
        prod.nombre        = form.nombre.data
        prod.descripcion   = form.descripcion.data
        prod.sku           = form.sku.data or None
        prod.codigo_barras = form.codigo_barras.data or None
        prod.categoria_id  = form.categoria_id.data
        prod.proveedor_id  = form.proveedor_id.data if form.proveedor_id.data else None
        prod.precio_compra = form.precio_compra.data or 0
        prod.precio_venta  = form.precio_venta.data
        prod.stock_minimo  = form.stock_minimo.data or 5
        prod.tallas  = json.dumps([t.strip() for t in form.tallas.data.split(',') if t.strip()]) if form.tallas.data else None
        prod.colores = json.dumps([c.strip() for c in form.colores.data.split(',') if c.strip()]) if form.colores.data else None
        prod.unidad  = form.unidad.data
        prod.activo  = form.activo.data
        if form.imagen.data and form.imagen.data.filename:
            try:
                prod.imagen_principal = _guardar_imagen(form.imagen.data)
            except ValueError as e:
                flash(str(e), 'danger')
                return render_template('admin/producto_form.html', form=form, titulo='Editar producto', producto=prod)
        db.session.commit()
        flash('Producto actualizado correctamente.', 'success')
        return redirect(url_for('admin.productos'))
    return render_template('admin/producto_form.html', form=form, titulo='Editar producto', producto=prod)


@bp.route('/productos/<int:id>')
@admin_requerido
def ver_producto(id):
    prod = Producto.query.get_or_404(id)
    movimientos = MovimientoStock.query.filter_by(producto_id=id)\
        .order_by(MovimientoStock.fecha.desc()).limit(20).all()
    return render_template('admin/producto_detalle.html', producto=prod, movimientos=movimientos)


@bp.route('/productos/<int:id>/toggle', methods=['POST'])
@admin_requerido
def toggle_producto(id):
    prod = Producto.query.get_or_404(id)
    prod.activo = not prod.activo
    db.session.commit()
    estado = 'activado' if prod.activo else 'desactivado'
    flash(f'Producto {estado}.', 'info')
    return redirect(url_for('admin.productos'))


# ═══════════════════════════════════════════════════════════════════════════════
# MOVIMIENTOS DE STOCK
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/movimientos')
@admin_requerido
def movimientos():
    pagina   = request.args.get('pagina', 1, type=int)
    tipo     = request.args.get('tipo', '')
    prod_id  = request.args.get('prod', 0, type=int)
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')

    query = MovimientoStock.query
    if tipo:
        query = query.filter_by(tipo=tipo)
    if prod_id:
        query = query.filter_by(producto_id=prod_id)
    if fecha_desde:
        try:
            query = query.filter(MovimientoStock.fecha >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
        except ValueError:
            pass
    if fecha_hasta:
        try:
            hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(MovimientoStock.fecha < hasta)
        except ValueError:
            pass

    movs      = query.order_by(MovimientoStock.fecha.desc()).paginate(page=pagina, per_page=25)
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    return render_template('admin/movimientos.html',
                           movimientos=movs, productos=productos,
                           tipo_sel=tipo, prod_sel=prod_id,
                           fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
                           TIPOS=MovimientoStock.TIPOS)


@bp.route('/movimientos/nuevo', methods=['GET', 'POST'])
@admin_requerido
def nuevo_movimiento():
    from forms.almacen_forms import FormMovimiento
    form = FormMovimiento()
    form.producto_id.choices = [(p.id, f'{p.nombre} (Stock: {p.stock})')
                                 for p in Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()]
    form.proveedor_id.choices = [(0, '— Sin proveedor —')] + [
        (p.id, p.nombre) for p in Proveedor.query.filter_by(activo=True).all()]
    if form.validate_on_submit():
        prod = Producto.query.get_or_404(form.producto_id.data)
        tipo = form.tipo.data
        cant = form.cantidad.data
        stock_antes = prod.stock

        if tipo in ('salida', 'venta'):
            if cant > prod.stock:
                flash(f'Stock insuficiente. Stock actual: {prod.stock}', 'danger')
                return render_template('admin/movimiento_form.html', form=form)
            prod.stock -= cant
        elif tipo in ('entrada', 'devolucion'):
            prod.stock += cant
        elif tipo == 'ajuste':
            prod.stock = cant   # En ajuste, la cantidad es el nuevo stock

        mov = MovimientoStock(
            producto_id=prod.id, tipo=tipo,
            cantidad=cant, stock_antes=stock_antes, stock_despues=prod.stock,
            motivo=form.motivo.data, referencia=form.referencia.data,
            proveedor_id=form.proveedor_id.data if form.proveedor_id.data else None,
            usuario_id=current_user.id,
        )
        db.session.add(mov)
        db.session.commit()
        flash('Movimiento registrado correctamente.', 'success')
        return redirect(url_for('admin.movimientos'))
    # Pre-seleccionar producto si viene por parámetro
    prod_id = request.args.get('prod', 0, type=int)
    if prod_id:
        form.producto_id.data = prod_id
    return render_template('admin/movimiento_form.html', form=form)


# ═══════════════════════════════════════════════════════════════════════════════
# VENTAS FÍSICAS
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/ventas')
@admin_requerido
def ventas():
    pagina      = request.args.get('pagina', 1, type=int)
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')
    metodo      = request.args.get('metodo', '')

    query = VentaFisica.query
    if not request.args.get('todas'):
        query = query.filter_by(anulada=False)
    if metodo:
        query = query.filter_by(metodo_pago=metodo)
    if fecha_desde:
        try:
            query = query.filter(VentaFisica.fecha >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
        except ValueError:
            pass
    if fecha_hasta:
        try:
            hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(VentaFisica.fecha < hasta)
        except ValueError:
            pass

    ventas_pag = query.order_by(VentaFisica.fecha.desc()).paginate(page=pagina, per_page=20)
    return render_template('admin/ventas.html', ventas=ventas_pag,
                           fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
                           metodo_sel=metodo, METODOS=VentaFisica.METODOS_PAGO)


@bp.route('/ventas/nueva', methods=['GET', 'POST'])
@admin_requerido
def nueva_venta():
    from forms.almacen_forms import FormVenta
    form = FormVenta()
    productos_activos = Producto.query.filter_by(activo=True)\
        .order_by(Producto.nombre).all()
    if request.method == 'POST':
        try:
            raw = request.form.get('items_json', '[]')
            print('DEBUG items_json:', repr(raw))
            items = json.loads(raw)
        except Exception:
            flash('Error al procesar los productos. Inténtalo de nuevo.', 'danger')
            return render_template('admin/venta_form.html', form=form,
                                   productos=productos_activos)
        if not items:
            flash('Agrega al menos un producto a la venta.', 'warning')
            return render_template('admin/venta_form.html', form=form,
                                   productos=productos_activos)

        descuento = float(form.descuento.data or 0)
        subtotal  = sum(float(i['subtotal']) for i in items)
        total     = max(subtotal - descuento, 0)

        venta = VentaFisica(
            numero_venta   = VentaFisica.generar_numero(),
            cliente_nombre = form.cliente_nombre.data or None,
            cliente_doc    = form.cliente_doc.data or None,
            metodo_pago    = form.metodo_pago.data,
            subtotal       = subtotal,
            descuento      = descuento,
            total          = total,
            notas          = form.notas.data,
            usuario_id     = current_user.id,
        )
        db.session.add(venta)
        db.session.flush()

        from collections import defaultdict
        cantidades_por_prod = defaultdict(int)
        for item in items:
            cantidades_por_prod[int(item['producto_id'])] += int(item['cantidad'])
        for prod_id_chk, cant_total in cantidades_por_prod.items():
            prod_chk = Producto.query.get(prod_id_chk)
            if not prod_chk or prod_chk.stock < cant_total:
                db.session.rollback()
                nombre_chk = prod_chk.nombre if prod_chk else f'ID {prod_id_chk}'
                flash(f'Stock insuficiente para: {nombre_chk}.', 'danger')
                return render_template('admin/venta_form.html', form=form,
                                       productos=productos_activos)
        for item in items:                        # ← mismo nivel que los for de arriba
            prod = Producto.query.get(item['producto_id'])
            if not prod or prod.stock < int(item['cantidad']):
                db.session.rollback()
                flash(f'Stock insuficiente para: {item.get("nombre","producto")}.', 'danger')
                return render_template('admin/venta_form.html', form=form,
                                       productos=productos_activos)
            detalle = DetalleVenta(
                venta_id    = venta.id,
                producto_id = prod.id,
                cantidad    = int(item['cantidad']),
                precio_unit = float(item['precio_unit']),
                talla       = item.get('talla') or None,
                color       = item.get('color') or None,
                subtotal    = float(item['subtotal']),
            )
            stock_antes = prod.stock
            prod.stock -= int(item['cantidad'])
            mov = MovimientoStock(
                producto_id  = prod.id, tipo='venta',
                cantidad     = int(item['cantidad']),
                stock_antes  = stock_antes, stock_despues=prod.stock,
                motivo       = f'Venta física {venta.numero_venta}',
                referencia   = venta.numero_venta,
                usuario_id   = current_user.id,
            )
            db.session.add(detalle)
            db.session.add(mov)

        db.session.commit()
        enviar_resumen_venta(venta)
        bajos = [d.producto for d in venta.detalles if d.producto.stock_bajo]
        if bajos:
            enviar_alerta_stock(bajos)
        flash(f'Venta {venta.numero_venta} registrada correctamente.', 'success')
        return redirect(url_for('admin.ver_venta', id=venta.id))

    return render_template('admin/venta_form.html', form=form,
                           productos=productos_activos)


@bp.route('/ventas/<int:id>')
@admin_requerido
def ver_venta(id):
    venta = VentaFisica.query.get_or_404(id)
    return render_template('admin/venta_detalle.html', venta=venta)


@bp.route('/ventas/<int:id>/anular', methods=['POST'])
@admin_requerido
def anular_venta(id):
    venta = VentaFisica.query.get_or_404(id)
    if venta.anulada:
        flash('Esta venta ya fue anulada.', 'warning')
        return redirect(url_for('admin.ver_venta', id=id))
    venta.anulada = True
    # Revertir stock
    for det in venta.detalles:
        prod = det.producto
        stock_antes = prod.stock
        prod.stock += det.cantidad
        mov = MovimientoStock(
            producto_id=prod.id, tipo='devolucion',
            cantidad=det.cantidad, stock_antes=stock_antes, stock_despues=prod.stock,
            motivo=f'Anulación venta {venta.numero_venta}',
            referencia=venta.numero_venta, usuario_id=current_user.id,
        )
        db.session.add(mov)
    db.session.commit()
    flash(f'Venta {venta.numero_venta} anulada y stock revertido.', 'info')
    return redirect(url_for('admin.ventas'))


# ═══════════════════════════════════════════════════════════════════════════════
# CATEGORÍAS
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/categorias')
@admin_requerido
def categorias():
    cats = Categoria.query.order_by(Categoria.nombre).all()
    return render_template('admin/categorias.html', categorias=cats)


@bp.route('/categorias/nueva', methods=['GET', 'POST'])
@admin_requerido
def nueva_categoria():
    form = FormCategoria()
    if form.validate_on_submit():
        slug = form.slug.data or _slugify(form.nombre.data)
        cat  = Categoria(nombre=form.nombre.data, slug=slug,
                         descripcion=form.descripcion.data,
                         icono=form.icono.data or 'box-seam',
                         activo=form.activo.data)
        db.session.add(cat)
        db.session.commit()
        flash('Categoría creada.', 'success')
        return redirect(url_for('admin.categorias'))
    return render_template('admin/categoria_form.html', form=form, titulo='Nueva categoría')


@bp.route('/categorias/<int:id>/editar', methods=['GET', 'POST'])
@admin_requerido
def editar_categoria(id):
    cat  = Categoria.query.get_or_404(id)
    form = FormCategoria(obj=cat)
    if form.validate_on_submit():
        cat.nombre      = form.nombre.data
        cat.slug        = form.slug.data or _slugify(form.nombre.data)
        cat.descripcion = form.descripcion.data
        cat.icono       = form.icono.data or 'box-seam'
        cat.activo      = form.activo.data
        db.session.commit()
        flash('Categoría actualizada.', 'success')
        return redirect(url_for('admin.categorias'))
    return render_template('admin/categoria_form.html', form=form,
                           titulo='Editar categoría', categoria=cat)


# ═══════════════════════════════════════════════════════════════════════════════
# PROVEEDORES
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/proveedores')
@admin_requerido
def proveedores():
    busqueda = request.args.get('q', '')
    query    = Proveedor.query
    if busqueda:
        query = query.filter(
            (Proveedor.nombre.ilike(f'%{busqueda}%')) |
            (Proveedor.ruc.ilike(f'%{busqueda}%'))
        )
    provs = query.order_by(Proveedor.nombre).all()
    return render_template('admin/proveedores.html', proveedores=provs, busqueda=busqueda)


@bp.route('/proveedores/nuevo', methods=['GET', 'POST'])
@admin_requerido
def nuevo_proveedor():
    form = FormProveedor()
    if form.validate_on_submit():
        prov = Proveedor(
            nombre=form.nombre.data, ruc=form.ruc.data or None,
            contacto=form.contacto.data, telefono=form.telefono.data,
            email=form.email.data, direccion=form.direccion.data,
            notas=form.notas.data, activo=form.activo.data,
        )
        db.session.add(prov)
        db.session.commit()
        flash('Proveedor registrado.', 'success')
        return redirect(url_for('admin.proveedores'))
    return render_template('admin/proveedor_form.html', form=form, titulo='Nuevo proveedor')


@bp.route('/proveedores/<int:id>/editar', methods=['GET', 'POST'])
@admin_requerido
def editar_proveedor(id):
    prov = Proveedor.query.get_or_404(id)
    form = FormProveedor(obj=prov)
    if form.validate_on_submit():
        prov.nombre    = form.nombre.data
        prov.ruc       = form.ruc.data or None
        prov.contacto  = form.contacto.data
        prov.telefono  = form.telefono.data
        prov.email     = form.email.data
        prov.direccion = form.direccion.data
        prov.notas     = form.notas.data
        prov.activo    = form.activo.data
        db.session.commit()
        flash('Proveedor actualizado.', 'success')
        return redirect(url_for('admin.proveedores'))
    return render_template('admin/proveedor_form.html', form=form,
                           titulo='Editar proveedor', proveedor=prov)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTES
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/reportes')
@admin_requerido
def reportes():
    hoy   = datetime.utcnow().date()
    mes_i = datetime(hoy.year, hoy.month, 1)

    # Ventas por mes (últimos 6 meses)
    ventas_por_mes = []
    for i in range(5, -1, -1):
        if hoy.month - i <= 0:
            m = hoy.month - i + 12
            y = hoy.year - 1
        else:
            m = hoy.month - i
            y = hoy.year
        desde = datetime(y, m, 1)
        hasta = datetime(y, m+1, 1) if m < 12 else datetime(y+1, 1, 1)
        total = db.session.query(func.sum(VentaFisica.total))\
            .filter(VentaFisica.fecha >= desde, VentaFisica.fecha < hasta,
                    VentaFisica.anulada==False).scalar() or 0
        ventas_por_mes.append({'mes': desde.strftime('%b %Y'), 'total': float(total)})

    # Ventas por método de pago (mes actual)
    ventas_metodo = db.session.query(
        VentaFisica.metodo_pago, func.sum(VentaFisica.total).label('total'),
        func.count(VentaFisica.id).label('cantidad')
    ).filter(VentaFisica.fecha >= mes_i, VentaFisica.anulada==False)\
     .group_by(VentaFisica.metodo_pago).all()

    # Top 10 productos más vendidos (mes actual)
    top_productos = db.session.query(
        Producto.nombre, Producto.sku,
        func.sum(DetalleVenta.cantidad).label('unidades'),
        func.sum(DetalleVenta.subtotal).label('monto')
    ).join(DetalleVenta, DetalleVenta.producto_id == Producto.id)\
     .join(VentaFisica, VentaFisica.id == DetalleVenta.venta_id)\
     .filter(VentaFisica.fecha >= mes_i, VentaFisica.anulada==False)\
     .group_by(Producto.id, Producto.nombre, Producto.sku)\
     .order_by(func.sum(DetalleVenta.subtotal).desc()).limit(10).all()

    # Inventario por categoría
    inventario_cat = db.session.query(
        Categoria.nombre,
        func.count(Producto.id).label('productos'),
        func.sum(Producto.stock).label('unidades'),
        func.sum(Producto.precio_venta * Producto.stock).label('valor')
    ).join(Producto, Producto.categoria_id == Categoria.id)\
     .filter(Producto.activo==True)\
     .group_by(Categoria.id, Categoria.nombre).all()

    # Productos sin stock / stock bajo
    sin_stock    = Producto.query.filter_by(activo=True).filter(Producto.stock <= 0).count()
    stock_bajo   = Producto.query.filter(Producto.stock > 0,
                                          Producto.stock <= Producto.stock_minimo,
                                          Producto.activo==True).count()

    return render_template('admin/reportes.html',
        ventas_por_mes=ventas_por_mes,
        ventas_metodo=ventas_metodo,
        top_productos=top_productos,
        inventario_cat=inventario_cat,
        sin_stock=sin_stock, stock_bajo=stock_bajo,
    )

# ═══════════════════════════════════════════════════════════════════════════════
# NOTIFICACIONES
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/enviar-resumen-diario', methods=['POST'])
@admin_requerido
def enviar_resumen_diario_manual():
    from utils.notificaciones import enviar_resumen_diario
    from sqlalchemy import func
    hoy = datetime.utcnow().date()

    ventas_hoy = VentaFisica.query\
        .filter(func.date(VentaFisica.fecha) == hoy, VentaFisica.anulada == False).all()

    total_hoy = sum(float(v.total) for v in ventas_hoy)

    productos_bajos = Producto.query.filter(
        Producto.stock <= Producto.stock_minimo,
        Producto.activo == True
    ).all()

    enviar_resumen_diario(ventas_hoy, total_hoy, productos_bajos)
    flash('Resumen del día enviado al correo.', 'success')
    return redirect(url_for('admin.dashboard'))

# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAR EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/exportar/productos')
@admin_requerido
def exportar_productos():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import Response
    import io

    productos = Producto.query.filter_by(activo=True)\
        .order_by(Producto.nombre).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'

    # Estilo encabezado
    header_font    = Font(bold=True, color='FFFFFF')
    header_fill    = PatternFill(fill_type='solid', fgColor='A53694')
    header_align   = Alignment(horizontal='center')

    encabezados = ['#', 'Nombre', 'SKU', 'Categoría', 'Proveedor',
                   'Precio Compra', 'Precio Venta', 'Stock', 'Stock Mínimo',
                   'Margen %', 'Valor Inventario', 'Estado']

    for col, titulo in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align

    # Filas
    for row, p in enumerate(productos, 2):
        ws.cell(row=row, column=1,  value=row - 1)
        ws.cell(row=row, column=2,  value=p.nombre)
        ws.cell(row=row, column=3,  value=p.sku or '—')
        ws.cell(row=row, column=4,  value=p.categoria.nombre)
        ws.cell(row=row, column=5,  value=p.proveedor.nombre if p.proveedor else '—')
        ws.cell(row=row, column=6,  value=float(p.precio_compra))
        ws.cell(row=row, column=7,  value=float(p.precio_venta))
        ws.cell(row=row, column=8,  value=p.stock)
        ws.cell(row=row, column=9,  value=p.stock_minimo)
        ws.cell(row=row, column=10, value=p.margen or 0)
        ws.cell(row=row, column=11, value=p.valor_inventario)
        ws.cell(row=row, column=12, value='Activo' if p.activo else 'Inactivo')

        # Colorear stock bajo en rojo
        if p.sin_stock:
            ws.cell(row=row, column=8).font = Font(bold=True, color='CC0000')
        elif p.stock_bajo:
            ws.cell(row=row, column=8).font = Font(bold=True, color='FF8800')

    # Ancho de columnas
    anchos = [4, 35, 15, 18, 20, 14, 14, 10, 12, 10, 16, 10]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = ancho

    # Guardar en memoria y enviar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fecha = datetime.utcnow().strftime('%Y%m%d')
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=productos_{fecha}.xlsx'}
    )


@bp.route('/exportar/ventas')
@admin_requerido
def exportar_ventas():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import Response
    import io

    ventas = VentaFisica.query.filter_by(anulada=False)\
        .order_by(VentaFisica.fecha.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill(fill_type='solid', fgColor='A53694')
    header_align = Alignment(horizontal='center')

    encabezados = ['#', 'N° Venta', 'Fecha', 'Cliente', 'Doc.',
                   'Método Pago', 'Subtotal', 'Descuento', 'Total', 'Notas']

    for col, titulo in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    total_general = 0
    for row, v in enumerate(ventas, 2):
        ws.cell(row=row, column=1,  value=row - 1)
        ws.cell(row=row, column=2,  value=v.numero_venta)
        ws.cell(row=row, column=3,  value=v.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=4,  value=v.cliente_nombre or '—')
        ws.cell(row=row, column=5,  value=v.cliente_doc or '—')
        ws.cell(row=row, column=6,  value=v.metodo_pago_label)
        ws.cell(row=row, column=7,  value=float(v.subtotal))
        ws.cell(row=row, column=8,  value=float(v.descuento))
        ws.cell(row=row, column=9,  value=float(v.total))
        ws.cell(row=row, column=10, value=v.notas or '')
        total_general += float(v.total)

    # Fila de total
    fila_total = len(ventas) + 2
    ws.cell(row=fila_total, column=8, value='TOTAL:').font = Font(bold=True)
    ws.cell(row=fila_total, column=9, value=total_general).font = Font(bold=True, color='A53694')

    anchos = [4, 22, 18, 25, 12, 16, 12, 12, 12, 30]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fecha = datetime.utcnow().strftime('%Y%m%d')
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=ventas_{fecha}.xlsx'}
    )

@bp.route('/exportar/movimientos')
@admin_requerido
def exportar_movimientos():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import Response
    import io

    movimientos = MovimientoStock.query.order_by(MovimientoStock.fecha.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Movimientos'

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill(fill_type='solid', fgColor='A53694')
    header_align = Alignment(horizontal='center')

    encabezados = ['#', 'Fecha', 'Tipo', 'Producto', 'Cantidad',
                   'Stock Ant.', 'Stock Nuevo', 'Motivo', 'Referencia', 'Proveedor', 'Usuario']

    for col, titulo in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for row, m in enumerate(movimientos, 2):
        ws.cell(row=row, column=1,  value=row - 1)
        ws.cell(row=row, column=2,  value=m.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=3,  value=m.tipo_label)
        ws.cell(row=row, column=4,  value=m.producto.nombre if m.producto else '—')
        ws.cell(row=row, column=5,  value=m.cantidad)
        ws.cell(row=row, column=6,  value=m.stock_antes)
        ws.cell(row=row, column=7,  value=m.stock_despues)
        ws.cell(row=row, column=8,  value=m.motivo or '—')
        ws.cell(row=row, column=9,  value=m.referencia or '—')
        ws.cell(row=row, column=10, value=m.proveedor.nombre if m.proveedor else '—')
        ws.cell(row=row, column=11, value=m.usuario.nombre if m.usuario else '—')

    anchos = [4, 18, 12, 30, 10, 10, 12, 35, 20, 25, 20]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fecha = datetime.utcnow().strftime('%Y%m%d')
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=movimientos_{fecha}.xlsx'}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# API AJAX
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/api/producto/<int:id>')
@admin_requerido
def api_producto(id):
    prod = Producto.query.get_or_404(id)
    return jsonify({
        'id': prod.id, 'nombre': prod.nombre,
        'precio_venta': float(prod.precio_venta),
        'stock': prod.stock, 'sku': prod.sku or '',
        'tallas': prod.tallas_lista, 'colores': prod.colores_lista,
        'unidad': prod.unidad,
    })


@bp.route('/api/buscar-productos')
@admin_requerido
def api_buscar_productos():
    q     = request.args.get('q', '')
    prods = Producto.query.filter(
        Producto.activo==True,
        (Producto.nombre.ilike(f'%{q}%')) | (Producto.sku.ilike(f'%{q}%'))
    ).limit(10).all()
    return jsonify([{
        'id': p.id, 'nombre': p.nombre,
        'precio_venta': float(p.precio_venta),
        'stock': p.stock, 'sku': p.sku or '',
    } for p in prods])